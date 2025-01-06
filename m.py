from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes
import openpyxl, os, hashlib, logging, time, threading, subprocess, sys

# مسیر فایل اکسل
EXCEL_FILE = "data.xlsx"

# توکن ربات
TOKEN = "8027936129:AAENv_C5K6e9eEg5XZdSYL2RD7AhLgrurCc"

# پیام خوشامدگویی
WELCOME_MESSAGE = (
    "همکاران عزیز و گرامی ☘️\n"
    "احتراماً لیست زیر، اسامی مجموعه‌هایی است که می‌بایست به آن‌ها خدمات فصلی ارائه گردد. خواهشمند است هر یک از شما عزیزان که خدمات فصلی را انجام می‌دهید، چه آن خدمات فصلی توسط برنامه‌ریزی برای شما تنظیم شده باشد و چه خودتان از این لیست برداشت کرده باشید، می‌بایست بر روی دکمه مربوط به آن پروژه کلیک کنید تا نام کاربری شما در کنار نام پروژه قرار گیرد. \n"
    "برای فراخوانی این لیست هر زمان که نیاز داشتید /fasly را ارسال نمایید. \n"
    "با سپاس از همراهی شما 🙏🌺"
)

# تنظیمات لاگ
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# تابع برای خواندن داده‌ها از اکسل
def load_data():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        ws.append(["Project", "Users"])
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Projects"]
    data = [(row[0].value, row[1].value if len(row) > 1 else None) for row in ws.iter_rows(min_row=2) if row[0].value]
    wb.close()
    return data

# تابع برای ذخیره داده‌ها در اکسل و انتقال انتخاب‌شده‌ها به انتها
def save_data(data):
    selected = [item for item in data if item[1]]  # پروژه‌های انتخاب‌شده
    unselected = [item for item in data if not item[1]]  # پروژه‌های انتخاب‌نشده
    sorted_data = unselected + selected  # انتقال انتخاب‌شده‌ها به انتهای لیست
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.append(["Project", "Users"])
    for project, users in sorted_data:
        ws.append([project, users])
    wb.save(EXCEL_FILE)

# ساخت callback_data ایمن
def generate_safe_callback_data(project_name):
    return hashlib.md5(project_name.encode()).hexdigest()[:10]

# ساخت کیبورد برای پروژه‌ها
def build_keyboard(page=0, items_per_page=15):
    data = load_data()
    keyboard = []
    start = page * items_per_page
    end = start + items_per_page
    for project, users in data[start:end]:
        emoji = "✅" if users else "🔵"
        user_display = f" ({users})" if users else ""
        safe_callback_data = generate_safe_callback_data(project)
        keyboard.append([InlineKeyboardButton(f"{project}{user_display} {emoji}", callback_data=safe_callback_data)])

    navigation_buttons = []
    if page > 0:
        navigation_buttons.append(InlineKeyboardButton("⬅️ صفحه قبل", callback_data=f"page_{page - 1}"))
    if end < len(data):
        navigation_buttons.append(InlineKeyboardButton("صفحه بعد ➡️", callback_data=f"page_{page + 1}"))

    if navigation_buttons:
        keyboard.append(navigation_buttons)

    return InlineKeyboardMarkup(keyboard)

# هندلر دستور /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(WELCOME_MESSAGE, reply_markup=build_keyboard())

# هندلر برای کلیک روی دکمه‌ها
async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data.startswith("page_"):
        page = int(query.data.split("_")[1])
        await query.edit_message_text(
            text=WELCOME_MESSAGE, reply_markup=build_keyboard(page=page)
        )
        return

    user = query.from_user.username or query.from_user.full_name
    project_clicked = query.data

    data = load_data()
    for i, (project, users) in enumerate(data):
        if generate_safe_callback_data(project) == project_clicked:
            users_list = set(users.split(", ")) if users else set()
            if user in users_list:
                users_list.remove(user)
            else:
                users_list.add(user)
            data[i] = (project, ", ".join(users_list) if users_list else None)
            break

    save_data(data)
    await query.edit_message_text(
        text=WELCOME_MESSAGE, reply_markup=build_keyboard()
    )

# هندلر برای دستور /fasly
async def fasly(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(WELCOME_MESSAGE, reply_markup=build_keyboard())

# مدیریت خطاها
async def error_handler(update, context):
    logger.error("Exception while handling an update:", exc_info=context.error)

# تابع اصلی برای اجرای ربات
def main():
    # ساخت برنامه
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("fasly", fasly))
    app.add_handler(CallbackQueryHandler(button))
    app.add_error_handler(error_handler)

    logger.info("ربات با موفقیت اجرا شد.")
    app.run_polling(poll_interval=1.0)

# تابع برای ریستارت هر دو دقیقه
def restart_every_two_minutes():
    while True:
        time.sleep(120)  # دو دقیقه
        logger.info("ریستارت خودکار فرآیند...")
        subprocess.Popen([sys.executable, os.path.abspath(__file__)])  # استفاده از مسیر کامل فایل و نسخه فعلی پایتون
        os._exit(0)  # بستن فرآیند فعلی

if __name__ == "__main__":
    restart_thread = threading.Thread(target=restart_every_two_minutes, daemon=True)
    restart_thread.start()
    main()
